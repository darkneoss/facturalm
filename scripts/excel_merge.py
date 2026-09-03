#!/usr/bin/env python3
"""Upsert idempotente de facturas en un .xlsx.

Esto es lo que ninguna de las apps de las que nace esta skill hace: ExtFact
regenera el Excel desde cero en cada corrida y solo *avisa* de los duplicados
del lote, asi que la fila duplicada se exporta igual. Aqui el Excel es
acumulativo y sobrevive entre corridas.

Clave de identidad: el UUID del Timbre Fiscal Digital. Cuando no hay UUID
(tipico de la ruta PDF), se usa una clave sustituta rfc|fecha|importe.

Uso:
    python excel_merge.py salida.xlsx --json datos.json
    cat datos.json | python excel_merge.py salida.xlsx --json -
"""
import argparse
import json
import os
import sys

from openpyxl import Workbook, load_workbook

import catalogos as cat

HOJA = "CFDI"
FORMATO_MONTO = "#,##0.00"


def clave(fila):
    """Identidad de la factura. UUID si lo hay; si no, clave sustituta."""
    uuid = (fila.get("uuid") or "").strip()
    if uuid and uuid.lower() not in ("sin timbre", "none", ""):
        return "uuid:" + uuid.upper()
    # Sin folio fiscal legible: emisor + fecha + monto es lo mas discriminante
    # que queda. Suficiente para no duplicar la misma factura re-procesada.
    return "sub:%s|%s|%s" % (
        (fila.get("rfc") or "").upper().strip(),
        (fila.get("fecha") or "").strip(),
        round(float(fila.get("importe") or 0), 2),
    )


def _encabezados(ws):
    return [c.value for c in ws[1]] if ws.max_row >= 1 else []


def _migrar_encabezados(ws):
    """Agrega al final las columnas de la skill que le falten a un Excel viejo.

    Sin esto, un Excel creado antes de que existiera una columna la perderia en
    silencio: el upsert solo escribe en las columnas que encuentra por nombre.
    Se agregan a la derecha para no recorrer las columnas propias del usuario.
    """
    presentes = {c.value for c in ws[1] if c.value}
    faltantes = [c for c in cat.COLUMNAS + cat.COLUMNAS_CONTROL if c not in presentes]
    for col in faltantes:
        ws.cell(row=1, column=ws.max_column + 1).value = col
    return faltantes


def _asegurar_hoja(ruta):
    """Abre el libro o lo crea con los encabezados. Devuelve (wb, ws, nuevo)."""
    if os.path.exists(ruta):
        wb = load_workbook(ruta)
        ws = wb[HOJA] if HOJA in wb.sheetnames else wb.create_sheet(HOJA)
        if ws.max_row < 1 or not _encabezados(ws) or _encabezados(ws)[0] is None:
            ws.append(cat.COLUMNAS + cat.COLUMNAS_CONTROL)
        else:
            _migrar_encabezados(ws)
        return wb, ws, False
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA
    ws.append(cat.COLUMNAS + cat.COLUMNAS_CONTROL)
    ws.freeze_panes = "A2"
    return wb, ws, True


def _indice_existente(ws, encabezados):
    """Mapa clave -> numero de fila, leyendo lo que ya hay en el Excel."""
    try:
        i_uuid = encabezados.index("UUID")
        i_rfc = encabezados.index("RFC")
        i_fecha = encabezados.index("Fecha")
        i_importe = encabezados.index("Importe")
    except ValueError:
        return {}
    indice = {}
    for n, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(fila):
            continue
        indice[clave({
            "uuid": fila[i_uuid], "rfc": fila[i_rfc],
            "fecha": fila[i_fecha], "importe": fila[i_importe] or 0,
        })] = n
    return indice


def fusionar(ruta_excel, filas, actualizar=False):
    """Agrega las filas nuevas. Devuelve conteos. No toca las columnas extra."""
    wb, ws, nuevo = _asegurar_hoja(ruta_excel)
    encabezados = _encabezados(ws)
    indice = _indice_existente(ws, encabezados)

    # Solo escribimos en las columnas que conocemos: si el usuario agrego
    # columnas propias a la derecha, se respetan (quedan vacias en las nuevas).
    posicion = {h: i for i, h in enumerate(encabezados) if h}
    i_numericas = [posicion[h] for h in cat.COLUMNAS_NUMERICAS if h in posicion]
    # Indices que son NUESTROS, por nombre y no por posicion: tras una
    # migracion las columnas de la skill pueden estar al final, despues de las
    # que agrego el usuario. Todo lo que no este aqui no se toca al actualizar.
    i_propias = {posicion[h] for h in cat.COLUMNAS + cat.COLUMNAS_CONTROL if h in posicion}

    n_agregadas = n_omitidas = n_actualizadas = n_revisar = n_error = 0

    for fila in filas:
        if fila.get("_confianza") == "error" or not fila.get("archivo"):
            n_error += 1
            continue

        valores = [None] * len(encabezados)
        for col, llave in zip(cat.COLUMNAS, cat.CLAVES):
            if col in posicion:
                valores[posicion[col]] = fila.get(llave)
        for col in cat.COLUMNAS_CONTROL:
            if col in posicion:
                valores[posicion[col]] = fila.get(col)

        k = clave(fila)
        if k in indice:
            if not actualizar:
                n_omitidas += 1
                continue
            destino = indice[k]
            for i, v in enumerate(valores):
                if i in i_propias:
                    ws.cell(row=destino, column=i + 1).value = v
            n_actualizadas += 1
        else:
            ws.append(valores)
            destino = ws.max_row
            indice[k] = destino
            n_agregadas += 1

        for i in i_numericas:
            ws.cell(row=destino, column=i + 1).number_format = FORMATO_MONTO
        if fila.get("_confianza") == "revisar":
            n_revisar += 1

    wb.save(ruta_excel)
    return {
        "excel": os.path.abspath(ruta_excel), "creado": nuevo,
        "agregadas": n_agregadas, "omitidas_duplicadas": n_omitidas,
        "actualizadas": n_actualizadas, "requieren_revision": n_revisar,
        "con_error": n_error, "total_en_excel": ws.max_row - 1,
    }


def main():
    p = argparse.ArgumentParser(description="Upsert idempotente de facturas en .xlsx")
    p.add_argument("excel")
    p.add_argument("--json", required=True, help="archivo JSON (objeto o array), o - para stdin")
    p.add_argument("--actualizar", action="store_true",
                   help="sobreescribir la fila si el UUID ya existe")
    args = p.parse_args()

    crudo = sys.stdin.read() if args.json == "-" else open(args.json, encoding="utf-8").read()
    datos = json.loads(crudo)
    if isinstance(datos, dict):
        datos = [datos]

    # Solo conteos por stdout: nunca contenido de facturas.
    json.dump(fusionar(args.excel, datos, args.actualizar), sys.stdout,
              ensure_ascii=False, indent=2)
    print()


if __name__ == "__main__":
    main()
