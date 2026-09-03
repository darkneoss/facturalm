#!/usr/bin/env python3
"""Comprueba la coherencia aritmetica del Excel de facturas.

La identidad que debe cumplir todo CFDI de ingreso:

    SubTotal - Descuento + Trasladados - Retenidos = Total

Sirve para detectar filas capturadas a mano desde un PDF con algun monto mal
leido. Reporta solo numeros de fila y diferencias: nunca vuelca datos de la
factura.

Uso:
    python verificar.py salida.xlsx
"""
import argparse
import sys

from openpyxl import load_workbook

import catalogos as cat

HOJA = "CFDI"
TOLERANCIA = 0.01   # centavos: el CFDI redondea a 2 decimales


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def verificar(ruta):
    wb = load_workbook(ruta, data_only=True)
    if HOJA not in wb.sheetnames:
        return None, "El libro no tiene la hoja '%s'." % HOJA
    ws = wb[HOJA]

    encabezados = [c.value for c in ws[1]]
    faltan = [c for c in ("Subtotal", "Importe", "Impuestos Trasladados",
                          "Impuestos Retenidos") if c not in encabezados]
    if faltan:
        return None, "Al Excel le faltan columnas: %s" % ", ".join(faltan)

    def col(nombre):
        return encabezados.index(nombre) if nombre in encabezados else None

    i_sub, i_tot = col("Subtotal"), col("Importe")
    i_tra, i_ret = col("Impuestos Trasladados"), col("Impuestos Retenidos")
    i_desc, i_org = col("Descuento"), col("_origen")
    sin_descuento = i_desc is None

    i_rfc, i_rfcr = col("RFC"), col("RFC Receptor")
    descuadres, rfc_malos, revisadas = [], [], 0
    for n, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(fila):
            continue
        revisadas += 1
        esperado = (_num(fila[i_sub])
                    - (_num(fila[i_desc]) if i_desc is not None else 0.0)
                    + _num(fila[i_tra]) - _num(fila[i_ret]))
        diff = round(esperado - _num(fila[i_tot]), 2)
        if abs(diff) > TOLERANCIA:
            descuadres.append({
                "fila": n, "diferencia": diff,
                "origen": fila[i_org] if i_org is not None else "?",
            })

        # Un RFC con digito verificador malo casi siempre es un error de OCR.
        # Importa doble: ensucia el dato y rompe la clave sustituta con la que
        # se deduplican las filas sin UUID.
        for etiqueta, idx in (("emisor", i_rfc), ("receptor", i_rfcr)):
            if idx is None:
                continue
            if cat.rfc_valido(fila[idx]) is False:
                rfc_malos.append({
                    "fila": n, "cual": etiqueta,
                    "origen": fila[i_org] if i_org is not None else "?",
                })

    return {
        "excel": ruta, "filas_revisadas": revisadas,
        "descuadres": descuadres, "rfc_invalidos": rfc_malos,
        "sin_columna_descuento": sin_descuento,
    }, None


def main():
    p = argparse.ArgumentParser(description="Verifica la aritmetica del Excel")
    p.add_argument("excel")
    args = p.parse_args()

    res, err = verificar(args.excel)
    if err:
        print("ERROR: %s" % err, file=sys.stderr)
        sys.exit(2)

    print("Filas revisadas: %d" % res["filas_revisadas"])
    if res["sin_columna_descuento"]:
        print("Aviso: el Excel no tiene columna Descuento; las facturas con "
              "descuento apareceran como descuadre. Vuelve a correr procesar.py "
              "para agregarla.")
    if res["rfc_invalidos"]:
        print("\n%d RFC con digito verificador invalido:" % len(res["rfc_invalidos"]))
        for r in res["rfc_invalidos"]:
            print("  fila %-4d RFC del %-9s (origen: %s)"
                  % (r["fila"], r["cual"], r["origen"]))
        print("Un RFC mal leido rompe la clave que evita duplicados. "
              "Corrigelo contra el documento.")

    if not res["descuadres"]:
        if not res["rfc_invalidos"]:
            print("OK: todas las filas cuadran.")
            return
        sys.exit(1)
    print("\n%d fila(s) no cuadran:" % len(res["descuadres"]))
    for d in res["descuadres"]:
        print("  fila %-4d diferencia %+10.2f  (origen: %s)"
              % (d["fila"], d["diferencia"], d["origen"]))
    print("\nLas de origen PDF suelen ser un monto mal leido; revisalas contra "
          "el documento.")
    sys.exit(1)


if __name__ == "__main__":
    main()
