#!/usr/bin/env python3
"""Verificacion de punta a punta con las facturas sinteticas.

No necesita datos reales: usa pruebas/facturas/, generado por generar.py.
Sale con codigo 1 si alguna comprobacion falla.

Uso: python pruebas/correr.py
"""
import json
import os
import shutil
import sys
import tempfile

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)
sys.path.insert(0, os.path.join(RAIZ, "scripts"))

import cfdi_xml          # noqa: E402
import pdf_texto         # noqa: E402
import excel_merge       # noqa: E402
import procesar          # noqa: E402
import verificar         # noqa: E402

FACTURAS = os.path.join(AQUI, "facturas")
fallos = []


def check(nombre, condicion, detalle=""):
    print("  %-46s %s" % (nombre, "OK" if condicion else "FALLA"))
    if not condicion:
        fallos.append("%s %s" % (nombre, detalle))


def main():
    if not os.path.isdir(FACTURAS) or not os.listdir(FACTURAS):
        print("Faltan las fixtures. Corre: python pruebas/generar.py")
        return 2

    tmp = tempfile.mkdtemp(prefix="facturalm-")
    try:
        xlsx = os.path.join(tmp, "prueba.xlsx")
        xmls, huerfanos = procesar.emparejar(FACTURAS)

        print("\nRuta XML")
        filas = [cfdi_xml.extraer(r) for r in xmls]
        r1 = excel_merge.fusionar(xlsx, filas)
        check("captura los 3 CFDI", r1["agregadas"] == 3, r1)

        r2 = excel_merge.fusionar(xlsx, filas)
        check("re-procesar no duplica", r2["agregadas"] == 0 and r2["omitidas_duplicadas"] == 3, r2)

        sin_timbre = [f for f in filas if f["estado"] == "Sin Timbre"]
        check("la factura sin timbre se captura marcada", len(sin_timbre) == 1)

        con_desc = [f for f in filas if f["descuento"] > 0]
        check("lee Comprobante@Descuento", len(con_desc) == 1)

        print("\nAritmetica")
        res, err = verificar.verificar(xlsx)
        check("Subtotal-Desc+Tras-Ret = Total", err is None and not res["descuadres"],
              err or res.get("descuadres"))

        print("\nRuta PDF (sin vision)")
        check("detecta el PDF sin XML", len(huerfanos) == 1)
        # El PDF de prueba tiene que llegar intacto. Si git lo trata como
        # texto y le inyecta CR al clonar, los offsets del xref se rompen y
        # pdf-inspector deja de abrirlo; pdfplumber si lo tolera, asi que sin
        # esta comprobacion el flujo cae al respaldo y la suite sigue en
        # verde con una fixture corrupta.
        if huerfanos and pdf_texto._pi is not None:
            check("la fixture PDF no llego corrupta",
                  pdf_texto.clasificar(huerfanos[0]) is not None,
                  "pdf-inspector no puede abrirla; revisa .gitattributes")
        if huerfanos:
            texto, meta = pdf_texto.extraer_texto(huerfanos[0])
            check("extrae texto plano", isinstance(texto, str) and len(texto) > 100)
            check("no usa OCR en un PDF con capa de texto", meta["metodo"] != "ocr", meta["metodo"])
            check("el texto conserva renglones", len(texto.splitlines()) >= 8,
                  len(texto.splitlines()))

        # Regresion: en facturas reales el texto posicionado rompe los acentos
        # (la 'e' acentuada sale como U+00D8, la 'o'/'i' se pierden) y el
        # clasificador no lo reporta. Si esta deteccion se rompe, los conceptos
        # llegan corruptos al Excel sin que nada avise.
        print("\nDeteccion de acentos rotos")
        roto = "Regimen fiscal del emisor. MOtodo de pago. CrOdito.".replace("O", "Ø")
        check("detecta la firma U+00D8 entre letras", pdf_texto.acentos_rotos(roto))
        sano = "Régimen fiscal del emisor. Método de pago. Crédito."
        check("no marca un texto sano", not pdf_texto.acentos_rotos(sano))
        # Una factura en MAYUSCULAS sin acentos es legitima: no debe dispararse
        # solo por no tener acentos, o desviaria a pdfplumber sin motivo.
        mayus = ("FACTURA CONSULTA MEDICA SUBTOTAL TOTAL IMPORTE CLIENTE " * 12)
        check("tolera texto legitimo sin acentos",
              not pdf_texto.acentos_rotos(mayus) or
              pdf_texto._acentos(mayus) == 0)
        check("cuenta acentos correctamente", pdf_texto._acentos(sano) == 3,
              pdf_texto._acentos(sano))

        # Un RFC mal leido por OCR no solo ensucia el dato: rompe la clave
        # rfc|fecha|importe con la que se deduplican las filas sin UUID.
        # Solo se usan RFC de prueba del SAT; nunca uno real.
        print("\nValidacion de RFC")
        import catalogos as _cat
        check("acepta el RFC de pruebas del SAT",
              _cat.rfc_valido("EKU9003173C9") is True)
        check("acepta los genericos (XAXX/XEXX)",
              _cat.rfc_valido("XAXX010101000") is True
              and _cat.rfc_valido("XEXX010101000") is True)
        check("rechaza un digito verificador alterado",
              _cat.rfc_valido("EKU9003173C8") is False)
        check("no juzga lo que no es un RFC",
              _cat.rfc_valido("Sin RFC") is None and _cat.rfc_valido("") is None)

        # La gente organiza las facturas por ano. Apuntar a la carpeta padre
        # tiene que funcionar, y dos archivos homonimos en anos distintos no
        # deben emparejarse entre si.
        print("\nCarpetas anidadas")
        anidado = os.path.join(tmp, "entrantes")
        for anio in ("2025", "2026"):
            os.makedirs(os.path.join(anidado, anio), exist_ok=True)
        shutil.copy(xmls[0], os.path.join(anidado, "2025", "A100.xml"))
        shutil.copy(huerfanos[0], os.path.join(anidado, "2026", "A100.pdf"))
        x2, h2 = procesar.emparejar(anidado)
        check("entra a las subcarpetas", len(x2) == 1 and len(h2) == 1,
              (len(x2), len(h2)))
        check("no empareja homonimos de anos distintos",
              h2 and h2[0].endswith("A100.pdf"))
        x3, h3 = procesar.emparejar(anidado, recursivo=False)
        check("--sin-recursion se queda en un nivel", not x3 and not h3)

        print("\nRuta OCR")
        if not pdf_texto.ruta_tesseract():
            print("  %-46s OMITIDA (Tesseract no instalado)" % "OCR de imagen")
        else:
            import re as _re

            import pypdfium2 as pdfium

            # Rasterizar el PDF de prueba simula una factura escaneada: la
            # imagen resultante no tiene capa de texto, solo pixeles.
            png = os.path.join(tmp, "escaneada.png")
            doc = pdfium.PdfDocument(huerfanos[0])
            doc[0].render(scale=2.0).to_pil().save(png)
            doc.close()

            texto, meta = pdf_texto.extraer_texto(png)
            check("una imagen suelta va por OCR", meta["metodo"] == "ocr-imagen",
                  meta["metodo"])
            check("el OCR recupera texto", len(texto) > 100, len(texto))
            check("el OCR recupera los RFC",
                  len(_re.findall(r"[A-Z]{3,4}[0-9]{6}[A-Z0-9]{3}", texto)) >= 2)
            if pdf_texto.idiomas_ocr() == "eng":
                print("  %-46s AVISO (falta el paquete 'spa')"
                      % "idioma espanol para OCR")

        print("\nErrores esperados")
        v33 = os.path.join(tmp, "v33.xml")
        with open(v33, "w", encoding="utf-8") as f:
            f.write('<?xml version="1.0"?><cfdi:Comprobante '
                    'xmlns:cfdi="http://www.sat.gob.mx/cfd/3" Version="3.3"/>')
        try:
            cfdi_xml.extraer(v33)
            check("rechaza CFDI 3.3", False)
        except cfdi_xml.CFDIError:
            check("rechaza CFDI 3.3", True)

        # Las facturas vienen de terceros. Un XML que declare una entidad
        # externa no debe poder leer archivos del disco: si esta prueba
        # empieza a fallar, el parser dejo de estar blindado.
        secreto = os.path.join(tmp, "secreto.txt")
        with open(secreto, "w", encoding="utf-8") as f:
            f.write("NO-DEBE-APARECER")
        xxe = os.path.join(tmp, "xxe.xml")
        with open(xxe, "w", encoding="utf-8") as f:
            f.write(
                '<?xml version="1.0"?>\n'
                '<!DOCTYPE c [<!ENTITY fuga SYSTEM "file:///%s">]>\n'
                '<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/4"'
                ' Version="4.0" SubTotal="1" Total="1">\n'
                '  <cfdi:Emisor Rfc="EKU9003173C9" Nombre="&fuga;"/>\n'
                '</cfdi:Comprobante>\n' % secreto.replace(os.sep, "/"))
        try:
            d = cfdi_xml.extraer(xxe)
            fuga = "NO-DEBE-APARECER" in json.dumps(d, ensure_ascii=False)
            check("no expande entidades externas (XXE)", not fuga)
        except cfdi_xml.CFDIError:
            check("no expande entidades externas (XXE)", True)

        print("\nExcel del usuario")
        from openpyxl import load_workbook
        wb = load_workbook(xlsx)
        ws = wb["CFDI"]
        ws.cell(row=1, column=ws.max_column + 1).value = "Notas"
        ws.cell(row=2, column=ws.max_column).value = "mia"
        wb.save(xlsx)
        excel_merge.fusionar(xlsx, filas, actualizar=True)
        ws = load_workbook(xlsx)["CFDI"]
        h = [c.value for c in ws[1]]
        check("preserva columnas propias al actualizar",
              "Notas" in h and ws.cell(row=2, column=h.index("Notas") + 1).value == "mia")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    print()
    if fallos:
        print("%d comprobacion(es) fallaron:" % len(fallos))
        for f in fallos:
            print("  - %s" % f)
        return 1
    print("Todo OK.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
